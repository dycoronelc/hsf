#!/usr/bin/env python3
"""Exporta docs/ubicacion_geo.xlsx (Cellbyte) → ubicacion_geo.csv en la raíz del repo.

Requiere: pip install openpyxl
Uso: python scripts/export-ubicacion-geo-from-xlsx.py
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "ubicacion_geo.xlsx"
OUT = ROOT / "ubicacion_geo.csv"
HEADER = [
    "PAIS",
    "PAISNAME",
    "PROVINCIA",
    "PROVINCIANAME",
    "DISTRITO",
    "DISTRITONAME",
    "CORREGCODE",
    "CORREGIMIENTO",
]


def cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def main() -> None:
    if not XLSX.exists():
        print(f"No existe {XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Excel vacío", file=sys.stderr)
        sys.exit(1)

    n = 0
    with OUT.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADER)
        for r in rows[1:]:
            if not r or r[2] is None:
                continue
            w.writerow([cell(r[i]) for i in range(8)])
            n += 1

    print(f"OK: {OUT} ({n} filas) desde {XLSX.name}")


if __name__ == "__main__":
    main()
